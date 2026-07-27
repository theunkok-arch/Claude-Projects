"""Storage: xlsx als master database + dagelijkse markdown digest.

- listings.xlsx bevat alle ooit gevonden listings (1 rij per listing, dedup via URL)
- digest/YYYY-MM-DD.md bevat de nieuwe listings van die dag, gesorteerd op score
"""
import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

XLSX_PATH = Path("listings.xlsx")
DIGEST_DIR = Path("digest")
SHEET_NAME = "Listings"

HEADERS = [
    "Datum gevonden",
    "Bron",
    "Titel",
    "Plaats",
    "Prijs",
    "m2",
    "Kamers",
    "URL",
    "Fit score",
    "AI motivatie",
    "Status",
]
URL_COL_IDX = HEADERS.index("URL")  # 0-indexed = 7


def _ensure_workbook() -> Workbook:
    """Open bestaande xlsx, of maak een nieuwe met headers."""
    if XLSX_PATH.exists():
        return load_workbook(XLSX_PATH)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)
    # Wat basis kolom-breedtes zodat het direct leesbaar is
    widths = [14, 14, 50, 18, 10, 6, 8, 60, 10, 60, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    wb.save(XLSX_PATH)
    return wb


def get_existing_urls() -> set:
    """Return set van URLs die al in de xlsx staan, voor dedup."""
    if not XLSX_PATH.exists():
        return set()
    try:
        wb = load_workbook(XLSX_PATH, read_only=True)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
        urls = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > URL_COL_IDX and row[URL_COL_IDX]:
                urls.add(row[URL_COL_IDX])
        return urls
    except Exception as e:
        print(f"[store] kon bestaande URLs niet lezen: {e}")
        return set()


def append_listings(scored_listings: list[dict]):
    """Voeg nieuwe rijen toe aan xlsx en schrijf dagelijkse markdown digest."""
    today = datetime.now().strftime("%Y-%m-%d")

    if not scored_listings:
        print("[store] geen nieuwe listings om toe te voegen")
        _write_digest(today, [])
        return

    wb = _ensure_workbook()
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    for l in scored_listings:
        ws.append([
            today,
            l.get("bron", ""),
            l.get("titel", ""),
            l.get("plaats", ""),
            l.get("prijs", ""),
            l.get("m2", ""),
            l.get("kamers", ""),
            l.get("url", ""),
            l.get("score", ""),
            l.get("motivatie", ""),
            "Nieuw",
        ])

    wb.save(XLSX_PATH)
    print(f"[store] {len(scored_listings)} listings toegevoegd aan {XLSX_PATH}")

    _write_digest(today, scored_listings)


def _write_digest(date_str: str, listings: list[dict]):
    """Schrijf digest/YYYY-MM-DD.md met de nieuwe listings van die dag."""
    DIGEST_DIR.mkdir(exist_ok=True)
    path = DIGEST_DIR / f"{date_str}.md"

    lines = [f"# Huurbot digest {date_str}", ""]

    if not listings:
        lines.append("_Geen nieuwe listings vandaag._")
        path.write_text("\n".join(lines), encoding="utf-8")
        print(f"[store] lege digest geschreven naar {path}")
        return

    lines.append(f"**{len(listings)} nieuwe listings** (gesorteerd op fit score)")
    lines.append("")
    lines.append("| Score | Titel | Plaats | Prijs | m2 | Kamers | Bron | Motivatie | Link |")
    lines.append("|---|---|---|---|---|---|---|---|---|")

    for l in listings:
        titel = (l.get("titel") or "").replace("|", "/")[:70]
        plaats = (l.get("plaats") or "").replace("|", "/")[:30]
        motivatie = (l.get("motivatie") or "").replace("|", "/")[:120]
        prijs = l.get("prijs") or ""
        m2 = l.get("m2") or ""
        kamers = l.get("kamers") or ""
        bron = l.get("bron") or ""
        score = l.get("score") or 0
        url = l.get("url") or ""
        link = f"[link]({url})" if url else ""
        lines.append(
            f"| {score} | {titel} | {plaats} | {prijs} | {m2} | {kamers} | {bron} | {motivatie} | {link} |"
        )

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"[store] digest geschreven naar {path}")
